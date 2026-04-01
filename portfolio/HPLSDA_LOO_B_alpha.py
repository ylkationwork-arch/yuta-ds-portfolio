import pandas as pd
import numpy as np
from sklearn.cross_decomposition import PLSRegression
from sklearn.metrics import roc_auc_score, accuracy_score, roc_curve
from sklearn.model_selection import LeaveOneOut
import matplotlib.pyplot as plt
import re
import os
from openpyxl import load_workbook
import string
import warnings
import unicodedata

# ==========================================
# 0. 設定
# ==========================================
warnings.filterwarnings("ignore")
plt.rcParams['font.family'] = 'MS Gothic'

TASKS_LIMIT = 30
MIN_SAMPLES = 3
N_ITERATIONS = 100 

CONV_THRESHOLD = 0.001 
STABLE_STREAK = 5 

output_dir = 'stage1_importance'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# ==========================================
# 1. ユーティリティ関数群
# ==========================================

def generate_excel_columns_upto_ZZ():
    cols = list(string.ascii_uppercase)
    for first in string.ascii_uppercase:
        for second in string.ascii_uppercase:
            cols.append(first + second)
    return cols

def colname_to_index(col):
    index = 0
    for c in col:
        index = index * 26 + ord(c.upper()) - ord('A') + 1
    return index - 1

def parse_cell_range(cell_range_str):
    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cell_range_str)
    if not match:
        raise ValueError(f"セル範囲の形式が正しくありません: '{cell_range_str}'")
    return colname_to_index(match[1]), int(match[2]), colname_to_index(match[3]), int(match[4])

def read_excel_data(file_path, sheet_index, cell_range_str):
    col_start, row_start, col_end, row_end = parse_cell_range(cell_range_str)
    df = pd.read_excel(file_path, sheet_name=sheet_index, header=None)
    if df.shape[1] <= col_end:
        for i in range(df.shape[1], col_end + 1):
            df[i] = np.nan
    return df.iloc[row_start - 1:row_end, col_start:col_end + 1]

# ==========================================
# 2. データ読み込みと整合
# ==========================================

alpab = generate_excel_columns_upto_ZZ()
print("結果出力先のセル番号 (例: J2):")
s = input()
match = re.match(r"([A-Z]+)(\d+)", s)
if match:
    alix = alpab.index(match.group(1))
    num = int(match.group(2))
else:
    alix = 0
    num = 1
    print("セル番号の指定が不正です。A1に出力します。")

config_file = "config.txt"
with open(config_file, 'r', encoding='utf-8') as file:
    lines = [line.strip() for line in file if line.strip()]

if len(lines) < 12:
    raise ValueError("config.txtには12行が必要です。")

raw_train_pos = read_excel_data(lines[0], int(lines[1]), lines[2])
raw_train_neg = read_excel_data(lines[3], int(lines[4]), lines[5])
raw_test_neg = read_excel_data(lines[6], int(lines[7]), lines[8])
raw_test_pos = read_excel_data(lines[9], int(lines[10]), lines[11])

def process_data(df):
    """入力データ自体のクリーニング"""
    df = df.reset_index(drop=True)
    
    # 1. 読み込み時点で inf を除去 (これは必須)
    df.replace([np.inf, -np.inf], np.nan, inplace=True)

    raw_tags = df.iloc[0, :].values
    raw_types = df.iloc[1, :].astype(str).values

    cleaned_types = []
    for t in raw_types:
        s = t
        if s.lower() in ['nan', 'none', '', 'null']:
            cleaned_types.append(np.nan)
        else:
            s = unicodedata.normalize('NFKC', s).strip()
            if s == "":
                cleaned_types.append(np.nan)
            else:
                cleaned_types.append(s)

    clinical = pd.DataFrame({'Tag': raw_tags, 'Type': cleaned_types})
    
    spectra = df.iloc[2:, :].T.apply(pd.to_numeric, errors='coerce')
    spectra.replace([np.inf, -np.inf], np.nan, inplace=True)
    spectra = spectra.fillna(0)
    
    return pd.concat([clinical.reset_index(drop=True), spectra.reset_index(drop=True)], axis=1)

def process_test_data_spectra_only(df):
    """テストデータ用クリーニング"""
    df_numeric = df.reset_index(drop=True).T.apply(pd.to_numeric, errors='coerce')
    df_numeric.replace([np.inf, -np.inf], np.nan, inplace=True)
    return df_numeric.fillna(0).reset_index(drop=True)

# データ処理実行
df_tr_p = process_data(raw_train_pos)
df_tr_n = process_data(raw_train_neg)

y_master = np.concatenate([np.ones(len(df_tr_p)), np.full(len(df_tr_n), -1)]).astype(int)
df_master = pd.concat([df_tr_p, df_tr_n], ignore_index=True)

X_test_neg_raw = process_test_data_spectra_only(raw_test_neg)
X_test_pos_raw = process_test_data_spectra_only(raw_test_pos)

y_ext_true = np.concatenate([np.full(len(X_test_neg_raw), -1), np.ones(len(X_test_pos_raw))]).astype(int)
df_ext_all = pd.concat([X_test_neg_raw, X_test_pos_raw], ignore_index=True)
df_ext_tags = pd.concat([
    pd.Series(raw_test_neg.iloc[0, :].values), 
    pd.Series(raw_test_pos.iloc[0, :].values)
], ignore_index=True)

# 波数整合
TAG_IDX, TYPE_IDX, SPECTRA_START_IDX = 0, 1, 2
with open("wavenumber1800-840.txt", 'r', encoding='utf-8') as f:
    wn_all = [float(l.strip()) for l in f if l.strip()]

valid_indices = [i for i, w in enumerate(wn_all) if 840 <= w <= 1800]

sel_cols_train = [SPECTRA_START_IDX + i for i in valid_indices if (SPECTRA_START_IDX + i) < df_master.shape[1]]
X_spec_master = df_master.iloc[:, sel_cols_train]
valid_wns = [wn_all[i] for i in valid_indices if (SPECTRA_START_IDX + i) < df_master.shape[1]]

X_spec_test_ext = df_ext_all.iloc[:, [i for i in valid_indices if i < df_ext_all.shape[1]]]

if X_spec_test_ext.shape[1] < len(valid_wns):
    pad = pd.DataFrame(0, index=X_spec_test_ext.index, columns=range(X_spec_test_ext.shape[1], len(valid_wns)))
    X_spec_test_ext = pd.concat([X_spec_test_ext, pad], axis=1)
X_spec_test_ext = X_spec_test_ext.iloc[:, :len(valid_wns)]

X_spec_master.columns = valid_wns
X_spec_test_ext.columns = valid_wns

# タスク生成
tasks = []
vals_h = df_master[y_master == 1]['Type'].dropna().unique()
vals_s = df_master[y_master == -1]['Type'].dropna().unique()

for vh in vals_h:
    for vs in vals_s:
        ch = len(df_master[(y_master==1) & (df_master['Type']==vh)])
        cs = len(df_master[(y_master==-1) & (df_master['Type']==vs)])
        if ch >= MIN_SAMPLES and cs >= MIN_SAMPLES:
            tasks.append({'tag': f"T_{vh}_vs_{vs}", 'vh': vh, 'vs': vs, 'priority': ch + cs})

tasks.sort(key=lambda x: x['priority'], reverse=True)
selected_tasks = tasks[:TASKS_LIMIT]

if len(selected_tasks) == 0:
    print("【警告】有効なタスクが生成されませんでした。Type列の内容を確認してください。")
else:
    print(f"生成されたタスク数: {len(selected_tasks)}")

# ==========================================
# 4. Stage 1: 個別課題モデル
# ==========================================
print("\n--- Stage 1: 個別課題学習 & 特徴抽出可視化 ---")
X_stage2 = pd.DataFrame(index=df_master.index)
X_stage2_test_ext = pd.DataFrame(index=X_spec_test_ext.index)

all_coefficients = []

if len(selected_tasks) > 0:
    for i, task in enumerate(selected_tasks):
        tag, vh, vs = task['tag'], task['vh'], task['vs']
        idx_in = df_master[(df_master['Type'] == vh) | (df_master['Type'] == vs)].index
        idx_out = df_master.index.difference(idx_in)
        
        scores_col = np.zeros(len(df_master))
        X_in, y_in = X_spec_master.loc[idx_in], y_master[idx_in]

        pls_full = PLSRegression(n_components=min(3, len(X_in)-1)).fit(X_in, y_in)
        coef = pls_full.coef_.ravel()
        all_coefficients.append(coef)

        if i < 10:
            plt.figure(figsize=(10, 4))
            plt.plot(valid_wns, coef, color='crimson', linewidth=1)
            safe_tag = re.sub(r'[\s\\/:*?"<>|]+', '_', str(tag))
            plt.title(f"PLS Coefficients - {tag}")
            plt.xlabel("Wavenumber (cm^-1)")
            plt.axhline(0, color='black', lw=0.8, ls='--')
            plt.gca().invert_xaxis()
            plt.grid(alpha=0.3)
            plt.tight_layout()
            plt.savefig(f'{output_dir}/{safe_tag}_coef.png')
            plt.close()

        for target_idx in idx_in:
            tr_loo = idx_in[idx_in != target_idx]
            pls_loo = PLSRegression(n_components=min(3, len(tr_loo)-1)).fit(X_spec_master.loc[tr_loo], y_master[tr_loo])
            pred = pls_loo.predict(X_spec_master.loc[[target_idx]]).ravel()[0]
            scores_col[df_master.index.get_loc(target_idx)] = pred

        if len(idx_out) > 0:
            pred_out = pls_full.predict(X_spec_master.loc[idx_out]).ravel()
            scores_col[df_master.index.get_indexer(idx_out)] = pred_out
                
        X_stage2[tag] = scores_col
        X_stage2_test_ext[tag] = pls_full.predict(X_spec_test_ext).ravel()
        
        if (i+1)%5==0: print(f"  Task {i+1}/{len(selected_tasks)} processed...")

    avg_importance = np.mean(np.abs(all_coefficients), axis=0)
    plt.figure(figsize=(12, 5))
    plt.fill_between(valid_wns, avg_importance, color='blue', alpha=0.5)
    plt.plot(valid_wns, avg_importance, color='blue', lw=1)
    plt.title("Stage 1 Overall Feature Importance")
    plt.xlabel("Wavenumber (cm^-1)")
    plt.gca().invert_xaxis()
    plt.tight_layout()
    plt.savefig(f'{output_dir}/_overall_importance.png')
    importance_df = pd.DataFrame({'Wavenumber': valid_wns, 'Mean_Abs_Coef': avg_importance})
    importance_df.to_csv("Stage1_Importance.csv", index=False)

# ==========================================
# 5. Stage 2: 最終統合 & 外部データ評価
# ==========================================
print("\n--- Stage 2: 最終判別モデル ---")

if X_stage2.shape[1] == 0:
    print("【エラー】有効なタスクがなく、Stage 2を実行できません。")
else:
    loo = LeaveOneOut()
    y_pred_cv = np.zeros(len(y_master))
    n_comp_s2 = min(5, X_stage2.shape[1])

    for tr_ix, te_ix in loo.split(X_stage2):
        pls_s2 = PLSRegression(n_components=n_comp_s2).fit(X_stage2.iloc[tr_ix], y_master[tr_ix])
        y_pred_cv[te_ix] = pls_s2.predict(X_stage2.iloc[te_ix]).ravel()[0]

    auc_cv = roc_auc_score((y_master==1).astype(int), y_pred_cv)
    acc_cv = accuracy_score(y_master, np.where(y_pred_cv >= 0, 1, -1))

    # 外部データの予測
    final_model = PLSRegression(n_components=n_comp_s2).fit(X_stage2, y_master)
    y_ext_raw_scores = final_model.predict(X_stage2_test_ext).ravel()

    # 5-3. ダウンサイジング + Youden指数
    print(f"\n--- External Validation: Downsizing & Youden Index ({N_ITERATIONS} iterations) ---")

    pos_indices = np.where(y_ext_true == 1)[0]
    neg_indices = np.where(y_ext_true == -1)[0]

    if len(pos_indices) == 0 or len(neg_indices) == 0:
        print("【エラー】外部データに両方のクラスが含まれていません。")
        acc_ext_total = 0; acc_ext_neg = 0; acc_ext_pos = 0
    else:
        optimal_thresholds = []
        acc_history = [] 
        
        for i in range(N_ITERATIONS):
            n_min = min(len(pos_indices), len(neg_indices))
            sel_pos = np.random.choice(pos_indices, n_min, replace=False)
            sel_neg = np.random.choice(neg_indices, n_min, replace=False)
            sel_indices = np.concatenate([sel_pos, sel_neg])
            
            y_iter_true = y_ext_true[sel_indices]
            y_iter_score = y_ext_raw_scores[sel_indices]
            
            # --- ここで無限大チェックを行う ---
            # もしサンプリングされたデータの中にinfが含まれていたら、その回の計算はスキップする
            if np.isinf(y_iter_score).any():
                continue # この回は無効としてスキップ

            fpr, tpr, thresholds = roc_curve(y_iter_true, y_iter_score, pos_label=1)
            j_scores = tpr - fpr 
            best_idx = np.argmax(j_scores)
            best_th = thresholds[best_idx]
            
            # 算出された閾値自体が有限かどうかもチェック
            if np.isfinite(best_th):
                optimal_thresholds.append(best_th)

                y_iter_pred = np.where(y_iter_score >= best_th, 1, -1)
                iter_acc = accuracy_score(y_iter_true, y_iter_pred)
                acc_history.append(iter_acc)
        
        # --- 最終的な閾値計算：有効な閾値のみの平均を取る ---
        if len(optimal_thresholds) > 0:
            final_threshold = np.mean(optimal_thresholds)
        else:
            print("【警告】すべての反復試行で数値エラーが発生しました。閾値を0とします。")
            final_threshold = 0.0
        
        # 収束判定
        if len(acc_history) > 0:
            acc_series = pd.Series(acc_history)
            cma_series = acc_series.expanding().mean()
            cma_diff = cma_series.diff().abs()
            is_stable = cma_diff < CONV_THRESHOLD
            stable_check = is_stable.rolling(window=STABLE_STREAK).sum()
            convergence_points = stable_check[stable_check == STABLE_STREAK].index
            
            print("-" * 30)
            if not convergence_points.empty:
                print(f"【収束判定】累積正解率は第 {int(convergence_points[0] - STABLE_STREAK + 2)} 回の有効試行で収束しました。")
            else:
                print(f"【収束判定】試行内では完全には収束しませんでした。")
            print(f" 最終累積正解率: {cma_series.iloc[-1]:.4f}")
        
        print(f" 平均最適閾値: {final_threshold:.4f}")
        print("-" * 30)

    # 5-4. 最終スコア調整 & Balanced Acc
    y_ext_adj_scores = y_ext_raw_scores - final_threshold
    y_ext_pred_class = np.where(y_ext_adj_scores >= 0, 1, -1)

    mask_neg = (y_ext_true == -1)
    mask_pos = (y_ext_true == 1)
    acc_ext_neg = accuracy_score(y_ext_true[mask_neg], y_ext_pred_class[mask_neg]) if np.sum(mask_neg) > 0 else 0
    acc_ext_pos = accuracy_score(y_ext_true[mask_pos], y_ext_pred_class[mask_pos]) if np.sum(mask_pos) > 0 else 0
    acc_ext_total = (acc_ext_neg + acc_ext_pos) / 2

    print("-" * 60)
    print(f"【学習データ CV】   Accuracy: {acc_cv:.4f}, AUC: {auc_cv:.4f}")
    print(f"【外部テストデータ】 Balanced Acc: {acc_ext_total:.4f}")
    print(f"                   Group(-1) Acc: {acc_ext_neg:.4f}")
    print(f"                   Group(+1) Acc: {acc_ext_pos:.4f}")
    print("-" * 60)

    try:
        wb = load_workbook('PLSDA_RESULT.xlsx')
        ws = wb['Accuracies']
        results = [
            "Hybrid-Filter-Inf", 
            round(auc_cv, 4), 
            round(acc_cv, 4), 
            "Ext-Bal-Acc", round(acc_ext_total, 4),
            "Ext-Neg", round(acc_ext_neg, 4),
            "Ext-Pos", round(acc_ext_pos, 4)
        ]
        for offset, val in enumerate(results):
            ws[f'{alpab[alix+offset]}{num}'] = val
        wb.save('PLSDA_RESULT.xlsx')
        print("Excelへの保存完了。")
    except Exception as e:
        print(f"Excel保存エラー: {e}")

    pd.DataFrame({'True':y_master, 'Score':y_pred_cv, 'Tag':df_master['Tag']}).to_csv("CV_Scores.csv", index=False, encoding='utf-8-sig')
    pd.DataFrame({
        'Tag': df_ext_tags,
        'True_Label': y_ext_true,
        'Raw_Score': y_ext_raw_scores,
        'Adj_Score': y_ext_adj_scores,
        'Pred': y_ext_pred_class
    }).to_csv("External_Scores_Adjusted.csv", index=False, encoding='utf-8-sig')

print(f"\nすべての処理が完了しました。")